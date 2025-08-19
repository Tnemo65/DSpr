#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV to Excel Converter with Advanced Formatting
===============================================

This script converts the 'done enriched.csv' file to a comprehensive Excel workbook
with multiple sheets, formatting, charts, and data analysis.

Author: Data Science Team
Date: August 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from datetime import datetime

class CSVToExcelConverter:
    """Advanced CSV to Excel converter with formatting and analysis"""
    
    def __init__(self, csv_file_path, output_file_path=None):
        self.csv_file_path = csv_file_path
        self.output_file_path = output_file_path or csv_file_path.replace('.csv', '_comprehensive.xlsx')
        self.data = None
        self.workbook = None
        
        # Define styling
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(horizontal="left", vertical="center")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def load_data(self):
        """Load CSV data with proper handling of different encodings"""
        print(f"Loading data from {self.csv_file_path}...")
        
        try:
            # Try UTF-8 first
            self.data = pd.read_csv(self.csv_file_path, encoding='utf-8')
        except UnicodeDecodeError:
            try:
                # Try UTF-8 with BOM
                self.data = pd.read_csv(self.csv_file_path, encoding='utf-8-sig')
            except UnicodeDecodeError:
                # Fall back to latin-1
                self.data = pd.read_csv(self.csv_file_path, encoding='latin-1')
        
        print(f"Data loaded successfully: {self.data.shape[0]} rows, {self.data.shape[1]} columns")
        
        # Clean column names
        self.data.columns = self.data.columns.str.strip()
        
        return self.data
    
    def create_workbook(self):
        """Create Excel workbook with multiple sheets"""
        print("Creating Excel workbook...")
        self.workbook = Workbook()
        
        # Remove default sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        
        return self.workbook
    
    def create_main_data_sheet(self):
        """Create main data sheet with all data and formatting"""
        print("Creating main data sheet...")
        
        ws = self.workbook.create_sheet("Complete Data")
        
        # Add data to worksheet
        for r in dataframe_to_rows(self.data, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create table
        if len(self.data) > 0:
            table = Table(displayName="DataTable", ref=f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}")
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        return ws
    
    def create_summary_sheet(self):
        """Create summary statistics sheet"""
        print("Creating summary sheet...")
        
        ws = self.workbook.create_sheet("Data Summary")
        
        # Basic info
        ws['A1'] = "Dataset Overview"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = "Total Records:"
        ws['B3'] = len(self.data)
        
        ws['A4'] = "Total Columns:"
        ws['B4'] = len(self.data.columns)
        
        ws['A5'] = "File Generated:"
        ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Column information
        ws['A7'] = "Column Information"
        ws['A7'].font = Font(bold=True, size=14)
        
        ws['A8'] = "Column Name"
        ws['B8'] = "Data Type"
        ws['C8'] = "Non-Null Count"
        ws['D8'] = "Null Count"
        ws['E8'] = "Null Percentage"
        
        # Style headers
        for col in ['A8', 'B8', 'C8', 'D8', 'E8']:
            ws[col].font = self.header_font
            ws[col].fill = self.header_fill
            ws[col].alignment = self.header_alignment
        
        # Add column info
        for i, col in enumerate(self.data.columns, start=9):
            ws[f'A{i}'] = col
            ws[f'B{i}'] = str(self.data[col].dtype)
            ws[f'C{i}'] = self.data[col].count()
            ws[f'D{i}'] = self.data[col].isnull().sum()
            ws[f'E{i}'] = f"{(self.data[col].isnull().sum() / len(self.data)) * 100:.2f}%"
        
        # Numeric columns summary
        numeric_cols = self.data.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            start_row = len(self.data.columns) + 11
            
            ws[f'A{start_row}'] = "Numeric Columns Statistics"
            ws[f'A{start_row}'].font = Font(bold=True, size=14)
            
            # Create summary for numeric columns
            numeric_summary = self.data[numeric_cols].describe()
            
            # Add headers
            headers = ['Statistic'] + list(numeric_cols)
            for i, header in enumerate(headers):
                cell = ws.cell(row=start_row + 2, column=i + 1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            
            # Add data
            for i, stat in enumerate(numeric_summary.index):
                ws.cell(row=start_row + 3 + i, column=1, value=stat)
                for j, col in enumerate(numeric_cols):
                    value = numeric_summary.loc[stat, col]
                    ws.cell(row=start_row + 3 + i, column=j + 2, value=f"{value:.2f}" if pd.notnull(value) else "N/A")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def create_charts_sheet(self):
        """Create charts and visualizations sheet"""
        print("Creating charts sheet...")
        
        ws = self.workbook.create_sheet("Data Visualizations")
        
        # Add title
        ws['A1'] = "Data Visualizations"
        ws['A1'].font = Font(bold=True, size=16)
        
        numeric_cols = self.data.select_dtypes(include=[np.number]).columns
        
        if len(numeric_cols) > 0:
            # Create a simple bar chart for numeric columns means
            chart_data = []
            for col in numeric_cols[:10]:  # Limit to first 10 numeric columns
                mean_val = self.data[col].mean()
                if pd.notnull(mean_val):
                    chart_data.append([col, mean_val])
            
            if chart_data:
                # Add data for chart
                ws['A3'] = "Column"
                ws['B3'] = "Average Value"
                
                for i, (col, val) in enumerate(chart_data, start=4):
                    ws[f'A{i}'] = col
                    ws[f'B{i}'] = val
                
                # Create bar chart
                chart = BarChart()
                chart.title = "Average Values by Numeric Columns"
                chart.x_axis.title = "Columns"
                chart.y_axis.title = "Average Value"
                
                data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(chart_data))
                cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(chart_data))
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                ws.add_chart(chart, "D3")
        
        return ws
    
    def create_filtered_sheets(self):
        """Create filtered sheets based on data characteristics"""
        print("Creating filtered data sheets...")
        
        # Try to identify different types of data and create separate sheets
        
        # If there are categorical columns, create value counts sheets
        categorical_cols = self.data.select_dtypes(include=['object']).columns
        
        for col in categorical_cols[:5]:  # Limit to first 5 categorical columns
            if len(self.data[col].unique()) <= 20:  # Only for columns with reasonable number of categories
                value_counts = self.data[col].value_counts()
                
                if len(value_counts) > 1:
                    sheet_name = f"{col[:25]}_Analysis"  # Limit sheet name length
                    ws = self.workbook.create_sheet(sheet_name)
                    
                    ws['A1'] = f"Analysis: {col}"
                    ws['A1'].font = Font(bold=True, size=14)
                    
                    ws['A3'] = "Value"
                    ws['B3'] = "Count"
                    ws['C3'] = "Percentage"
                    
                    # Style headers
                    for cell_ref in ['A3', 'B3', 'C3']:
                        ws[cell_ref].font = self.header_font
                        ws[cell_ref].fill = self.header_fill
                        ws[cell_ref].alignment = self.header_alignment
                    
                    # Add data
                    total = len(self.data)
                    for i, (value, count) in enumerate(value_counts.items(), start=4):
                        ws[f'A{i}'] = str(value)
                        ws[f'B{i}'] = count
                        ws[f'C{i}'] = f"{(count/total)*100:.2f}%"
                    
                    # Create pie chart if reasonable number of categories
                    if len(value_counts) <= 10:
                        chart = PieChart()
                        chart.title = f"Distribution of {col}"
                        
                        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(value_counts))
                        labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(value_counts))
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        
                        ws.add_chart(chart, "E3")
    
    def finalize_workbook(self):
        """Finalize workbook with additional formatting and protection"""
        print("Finalizing workbook...")
        
        # Set active sheet to the first one
        if len(self.workbook.worksheets) > 0:
            self.workbook.active = self.workbook.worksheets[0]
        
        # Add workbook properties
        self.workbook.properties.title = "Data Analysis Report"
        self.workbook.properties.subject = "Comprehensive data analysis from CSV"
        self.workbook.properties.creator = "CSV to Excel Converter"
        self.workbook.properties.description = f"Generated from {os.path.basename(self.csv_file_path)} on {datetime.now().strftime('%Y-%m-%d')}"
    
    def convert(self):
        """Main conversion method"""
        print("=" * 60)
        print("CSV TO EXCEL COMPREHENSIVE CONVERTER")
        print("=" * 60)
        
        try:
            # Load data
            self.load_data()
            
            # Create workbook
            self.create_workbook()
            
            # Create sheets
            self.create_main_data_sheet()
            self.create_summary_sheet()
            self.create_charts_sheet()
            self.create_filtered_sheets()
            
            # Finalize
            self.finalize_workbook()
            
            # Save workbook
            print(f"Saving workbook to {self.output_file_path}...")
            self.workbook.save(self.output_file_path)
            
            print("=" * 60)
            print("CONVERSION COMPLETED SUCCESSFULLY!")
            print(f"Input file: {self.csv_file_path}")
            print(f"Output file: {self.output_file_path}")
            print(f"Number of sheets created: {len(self.workbook.worksheets)}")
            print(f"Total records processed: {len(self.data):,}")
            print("=" * 60)
            
            return True
            
        except Exception as e:
            print(f"Error during conversion: {str(e)}")
            return False

def main():
    """Main function to run the converter"""
    
    # File paths
    input_file = "done enriched.csv"
    output_file = "done_enriched_comprehensive.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found!")
        print("Please make sure the file exists in the current directory.")
        return
    
    # Create converter instance
    converter = CSVToExcelConverter(input_file, output_file)
    
    # Run conversion
    success = converter.convert()
    
    if success:
        print(f"\n✅ Success! Your Excel file has been created: {output_file}")
        print("\nThe Excel file includes:")
        print("  📊 Complete Data - All original data with formatting")
        print("  📈 Data Summary - Statistics and column information")
        print("  📉 Data Visualizations - Charts and graphs")
        print("  🔍 Analysis Sheets - Filtered views for categorical data")
        print("\nYou can now open the file in Excel for further analysis!")
    else:
        print("\n❌ Conversion failed. Please check the error messages above.")

if __name__ == "__main__":
    main()
